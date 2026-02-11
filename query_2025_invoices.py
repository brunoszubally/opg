"""
Script to query 2025 invoices from NAV Online Invoice system
Filter by DELIVERY DATE (teljesítési dátum) and export to Excel
"""

from datetime import datetime, timedelta
import json
from nav_online_invoice import (
    NavOnlineInvoiceConfig,
    NavOnlineInvoiceReporter,
)

# User credentials
user_data = {
    "login": "lyiql72hnvnaa30",
    "password": "Vilaguralom1472!",
    "taxNumber": "90043589",
    "signKey": "31-bc22-c0e9ce8cc07255DXRST0SUI3",
    "exchangeKey": "f6f655DXRST0RTTU"
}

# Software data (fixed values)
software_data = {
    "softwareId": "123456789123456789",
    "softwareName": "string",
    "softwareOperation": "ONLINE_SERVICE",
    "softwareMainVersion": "string",
    "softwareDevName": "string",
    "softwareDevContact": "string",
    "softwareDevCountryCode": "HU",
    "softwareDevTaxNumber": "string"
}

def query_all_invoices_paginated(reporter, invoice_query_params, max_pages=100):
    """Query all invoices with pagination"""
    page = 1
    all_invoices = []
    
    while page <= max_pages:
        print(f"  Querying page {page}...")
        
        result = reporter.query_invoice_digest(invoice_query_params, page, "OUTBOUND")
        
        invoices = result.get('invoiceDigest', [])
        if invoices:
            print(f"  Found {len(invoices)} invoices on page {page}")
            all_invoices.extend(invoices)
        
        available_pages = result.get('availablePage', 1)
        
        if page >= available_pages:
            break
        
        page += 1
    
    return all_invoices

def filter_by_delivery_date(invoices, year):
    """Filter invoices by delivery date year"""
    filtered = []
    for inv in invoices:
        delivery_date = inv.get('invoiceDeliveryDate', '')
        if delivery_date and delivery_date.startswith(str(year)):
            filtered.append(inv)
    return filtered

def export_to_excel(invoices, filename):
    """Export invoices to Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.run(['pip', 'install', 'openpyxl'], check=True)
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Számlák 2025"
    
    # Headers
    headers = [
        "Számla szám",
        "Kiállítás dátuma",
        "Teljesítés dátuma",
        "Partner név",
        "Partner adószám",
        "Nettó összeg (HUF)",
        "ÁFA összeg (HUF)",
        "Bruttó összeg (HUF)",
        "Művelet",
        "Számla kategória"
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Write data
    for row, inv in enumerate(invoices, 2):
        net_amount = float(inv.get('invoiceNetAmountHUF', 0))
        vat_amount = float(inv.get('invoiceVatAmountHUF', 0))
        gross_amount = net_amount + vat_amount
        
        data = [
            inv.get('invoiceNumber', ''),
            inv.get('invoiceIssueDate', ''),
            inv.get('invoiceDeliveryDate', ''),
            inv.get('supplierName', inv.get('customerName', '')),
            inv.get('supplierTaxNumber', inv.get('customerTaxNumber', '')),
            net_amount,
            vat_amount,
            gross_amount,
            inv.get('invoiceOperation', ''),
            inv.get('invoiceCategory', '')
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col in [6, 7, 8]:  # Number columns
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
    
    # Adjust column widths
    column_widths = [15, 15, 15, 30, 15, 18, 18, 18, 12, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(filename)
    print(f"\nExcel file saved: {filename}")

def main():
    print("Creating NAV Online Invoice client...")
    config = NavOnlineInvoiceConfig(NavOnlineInvoiceConfig.PROD_URL, user_data, software_data)
    reporter = NavOnlineInvoiceReporter(config)
    print("NAV client created successfully")
    
    # We need to query by issue date, then filter by delivery date
    # Query a wider range (2024-2026) to catch all invoices with 2025 delivery dates
    all_invoices = []
    
    # Query 2024 Q4 (might have 2025 delivery dates)
    print("\n=== Querying 2024 Q4 (for possible 2025 delivery dates) ===")
    for month in range(10, 13):
        date_from = f"2024-{month:02d}-01"
        if month == 12:
            date_to = "2024-12-31"
        else:
            next_month = datetime(2024, month + 1, 1)
            last_day = (next_month - timedelta(days=1)).day
            date_to = f"2024-{month:02d}-{last_day:02d}"
        
        print(f"\n  Month 2024-{month}: {date_from} to {date_to}")
        
        invoice_query_params = {
            "mandatoryQueryParams": {
                "invoiceIssueDate": {
                    "dateFrom": date_from,
                    "dateTo": date_to,
                },
            },
        }
        
        try:
            monthly_invoices = query_all_invoices_paginated(reporter, invoice_query_params)
            print(f"  Total: {len(monthly_invoices)} invoices")
            all_invoices.extend(monthly_invoices)
        except Exception as ex:
            print(f"  Error: {ex}")
    
    # Query all of 2025
    print("\n=== Querying 2025 (full year) ===")
    for month in range(1, 13):
        date_from = f"2025-{month:02d}-01"
        if month == 12:
            date_to = "2025-12-31"
        else:
            next_month = datetime(2025, month + 1, 1)
            last_day = (next_month - timedelta(days=1)).day
            date_to = f"2025-{month:02d}-{last_day:02d}"
        
        print(f"\n  Month 2025-{month}: {date_from} to {date_to}")
        
        invoice_query_params = {
            "mandatoryQueryParams": {
                "invoiceIssueDate": {
                    "dateFrom": date_from,
                    "dateTo": date_to,
                },
            },
        }
        
        try:
            monthly_invoices = query_all_invoices_paginated(reporter, invoice_query_params)
            print(f"  Total: {len(monthly_invoices)} invoices")
            all_invoices.extend(monthly_invoices)
        except Exception as ex:
            print(f"  Error: {ex}")
    
    # Query 2026 Q1 (might have 2025 delivery dates)
    print("\n=== Querying 2026 Q1 (for possible 2025 delivery dates) ===")
    for month in range(1, 3):
        date_from = f"2026-{month:02d}-01"
        if month == 2:
            date_to = "2026-02-09"  # Current date
        else:
            date_to = "2026-01-31"
        
        print(f"\n  Month 2026-{month}: {date_from} to {date_to}")
        
        invoice_query_params = {
            "mandatoryQueryParams": {
                "invoiceIssueDate": {
                    "dateFrom": date_from,
                    "dateTo": date_to,
                },
            },
        }
        
        try:
            monthly_invoices = query_all_invoices_paginated(reporter, invoice_query_params)
            print(f"  Total: {len(monthly_invoices)} invoices")
            all_invoices.extend(monthly_invoices)
        except Exception as ex:
            print(f"  Error: {ex}")
    
    print(f"\n{'='*60}")
    print(f"TOTAL INVOICES FETCHED: {len(all_invoices)}")
    
    # Filter by 2025 delivery date
    filtered_invoices = filter_by_delivery_date(all_invoices, 2025)
    print(f"INVOICES WITH 2025 DELIVERY DATE: {len(filtered_invoices)}")
    print(f"{'='*60}")
    
    # Calculate summary
    total_amount = 0.0
    storno_count = 0
    create_count = 0
    modify_count = 0
    
    for inv in filtered_invoices:
        amount = float(inv.get('invoiceNetAmountHUF', 0))
        operation = inv.get('invoiceOperation', 'CREATE')
        
        if operation == 'STORNO':
            storno_count += 1
            total_amount -= abs(amount)
        elif operation == 'CREATE':
            create_count += 1
            total_amount += amount
        elif operation == 'MODIFY':
            modify_count += 1
            total_amount += amount
    
    print(f"\nSUMMARY (2025 delivery date):")
    print(f"  CREATE invoices: {create_count}")
    print(f"  STORNO invoices: {storno_count}")
    print(f"  MODIFY invoices: {modify_count}")
    print(f"  Net amount (HUF): {total_amount:,.0f}")
    
    # Export to Excel
    tax_number = user_data['taxNumber']
    filename = f"szamlak_{tax_number}_2025_teljesites.xlsx"
    export_to_excel(filtered_invoices, filename)

if __name__ == "__main__":
    main()
